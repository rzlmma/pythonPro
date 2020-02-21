# -*- coding:utf-8 -*-
# __author__ = majing
import os
import sys

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.append(BASE_DIR)

import pandas as pd
from openpyxl import load_workbook

file_name = u"2020年度国家公务员考试招考简章.xlsx"


def deal_excel(file_name):
    rest = []
    wb = load_workbook(file_name)
    sheet_names = wb.sheetnames
    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows(min_row=2, max_col=13):
            subject = row[12].value

            if not subject:
                continue

            if '；' in subject:
                for item in subject.split('；'):
                    if '：' in item:
                        s__1 = item.split('：')[1]
                        s_list = s__1.split('、')
                        rest.extend(s_list)
    rest = list(set(rest))

    print("rest:\n %s" % rest)

if __name__ == "__main__":
    deal_excel(file_name)


