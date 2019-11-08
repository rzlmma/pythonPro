# -*- coding:utf-8 -*-
# __author__ = majing

from openpyxl import load_workbook
from collections import defaultdict

exam_forms = {'读':[4,5,3,78], '说':[8,9,64,70], '听':[8,9,3,63], '写':[4,5,79,79]}


def read_excel(file_name, sheet_p):
    data = defaultdict(list)
    wb = load_workbook(file_name)
    sheet = wb['Sheet1']

    for item in sheet.iter_rows(min_col=sheet_p[0], max_col=sheet_p[1], min_row=sheet_p[2], max_row=sheet_p[3]):
        first, second = item
        data[first.value.strip()].append(second.value.strip())
    return data


def read_excel_1(file_name):
    data = defaultdict(list)
    wb = load_workbook(file_name)
    sheet = wb['Sheet1']

    for item in sheet.iter_rows(min_row=1):
        first, second = item
        data[first.value.strip()].append(second.value.strip())
    return data



if __name__ == "__main__":
   data = read_excel_1("yy_upload_papers.xlsx")
   print("data: ", data)